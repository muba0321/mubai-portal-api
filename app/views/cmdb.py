from flask import Blueprint, request, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.extensions import db
from app.models.cmdb_vm import CmdbVM
from app.utils.response import success, error, page_result
import io
import re

cmdb_bp = Blueprint("cmdb", __name__)


def vm_to_dict(vm):
    return {
        "id": vm.id,
        "name": vm.name,
        "cluster": vm.cluster,
        "external_ip": vm.external_ip,
        "internal_ip": vm.internal_ip,
        "description": vm.description,
        "status": vm.status,
        "tenant": vm.tenant,
        "vcpus": vm.vcpus,
        "memory": vm.memory,
        "disk": vm.disk,
        "access_url": vm.access_url,
        "created_at": vm.created_at,
        "updated_at": vm.updated_at,
    }


@cmdb_bp.route("/vms", methods=["GET"])
@jwt_required()
def list_vms():
    keywords = request.args.get("keywords", "")
    cluster = request.args.get("cluster", "")
    status = request.args.get("status")
    tenant = request.args.get("tenant", "")
    page_num = int(request.args.get("pageNum", 1))
    page_size = int(request.args.get("pageSize", 10))

    q = CmdbVM.query.filter_by(deleted=0)

    if keywords:
        kw = f"%{keywords}%"
        q = q.filter(
            db.or_(
                CmdbVM.name.like(kw),
                CmdbVM.cluster.like(kw),
                CmdbVM.description.like(kw),
                CmdbVM.external_ip.like(kw),
                CmdbVM.internal_ip.like(kw),
            )
        )
    if cluster:
        q = q.filter_by(cluster=cluster)
    if status is not None:
        q = q.filter_by(status=int(status))
    if tenant:
        q = q.filter_by(tenant=tenant)

    q = q.order_by(CmdbVM.updated_at.desc())
    pagination = q.paginate(page=page_num, per_page=page_size, error_out=False)

    items = [vm_to_dict(vm) for vm in pagination.items]
    return success(data=page_result(pagination.total, items))


@cmdb_bp.route("/vms/<int:vm_id>", methods=["GET"])
@jwt_required()
def get_vm(vm_id):
    vm = CmdbVM.query.filter_by(id=vm_id, deleted=0).first()
    if not vm:
        return error(msg="虚拟机不存在", code="40400")
    return success(data=vm_to_dict(vm))


@cmdb_bp.route("/vms", methods=["POST"])
@jwt_required()
def create_vm():
    data = request.get_json()
    if not data.get("name") or not data.get("cluster") or not data.get("externalIp") or not data.get("internalIp") or not data.get("tenant"):
        return error(msg="缺少必填字段")

    if CmdbVM.query.filter_by(name=data["name"], deleted=0).first():
        return error(msg="虚拟机名称已存在")

    vm = CmdbVM(
        name=data["name"],
        cluster=data["cluster"],
        external_ip=data["externalIp"],
        internal_ip=data["internalIp"],
        description=data.get("description", ""),
        status=data.get("status", 1),
        tenant=data["tenant"],
        vcpus=data.get("vcpus", 4),
        memory=data.get("memory", 8192),
        disk=data.get("disk", ""),
        access_url=data.get("accessUrl", ""),
        created_by=get_jwt_identity(),
    )
    db.session.add(vm)
    db.session.commit()
    return success(msg="新增成功")


@cmdb_bp.route("/vms/<int:vm_id>", methods=["PUT"])
@jwt_required()
def update_vm(vm_id):
    vm = CmdbVM.query.filter_by(id=vm_id, deleted=0).first()
    if not vm:
        return error(msg="虚拟机不存在", code="40400")

    data = request.get_json()
    if "name" in data:
        existing = CmdbVM.query.filter_by(name=data["name"], deleted=0).first()
        if existing and existing.id != vm_id:
            return error(msg="虚拟机名称已存在")
        vm.name = data["name"]
    if "cluster" in data:
        vm.cluster = data["cluster"]
    if "externalIp" in data:
        vm.external_ip = data["externalIp"]
    if "internalIp" in data:
        vm.internal_ip = data["internalIp"]
    if "description" in data:
        vm.description = data["description"]
    if "status" in data:
        vm.status = data["status"]
    if "tenant" in data:
        vm.tenant = data["tenant"]
    if "vcpus" in data:
        vm.vcpus = data["vcpus"]
    if "memory" in data:
        vm.memory = data["memory"]
    if "disk" in data:
        vm.disk = data["disk"]
    if "accessUrl" in data:
        vm.access_url = data["accessUrl"]
    vm.updated_by = get_jwt_identity()

    db.session.commit()
    return success(msg="修改成功")


@cmdb_bp.route("/vms/<int:vm_id>", methods=["DELETE"])
@jwt_required()
def delete_vm(vm_id):
    vm = CmdbVM.query.filter_by(id=vm_id, deleted=0).first()
    if not vm:
        return error(msg="虚拟机不存在", code="40400")
    vm.deleted = 1
    vm.updated_by = get_jwt_identity()
    db.session.commit()
    return success(msg="删除成功")


@cmdb_bp.route("/vms/batch", methods=["DELETE"])
@jwt_required()
def batch_delete_vms():
    data = request.get_json()
    ids = data.get("ids", [])
    if not ids:
        return error(msg="请选择要删除的虚拟机")

    CmdbVM.query.filter(CmdbVM.id.in_(ids), CmdbVM.deleted == 0).update(
        {"deleted": 1}, synchronize_session=False
    )
    db.session.commit()
    return success(msg="批量删除成功")


@cmdb_bp.route("/vms/import", methods=["POST"])
@jwt_required()
def import_vms():
    if "file" not in request.files:
        return error(msg="请上传文件")

    file = request.files["file"]
    filename = file.filename.lower()

    try:
        if filename.endswith(".csv"):
            content = file.read().decode("utf-8-sig")
            return _import_csv(content)
        elif filename.endswith((".xlsx", ".xls")):
            return _import_excel(file)
        else:
            return error(msg="仅支持 Excel/CSV 文件")
    except Exception as e:
        return error(msg=f"导入失败: {str(e)}")


def _import_csv(content):
    lines = content.strip().split("\n")
    if len(lines) < 2:
        return error(msg="文件内容为空")

    success_count = 0
    fail_count = 0
    errors = []

    for i, line in enumerate(lines[1:], start=2):
        parts = line.strip().split(",")
        if len(parts) < 4:
            fail_count += 1
            errors.append(f"第{i}行：数据不完整")
            continue

        name = parts[0].strip()
        cluster = parts[1].strip()
        external_ip = parts[2].strip()
        internal_ip = parts[3].strip()

        if CmdbVM.query.filter_by(name=name, deleted=0).first():
            fail_count += 1
            errors.append(f"第{i}行：名称 '{name}' 已存在")
            continue

        vm = CmdbVM(
            name=name, cluster=cluster, external_ip=external_ip,
            internal_ip=internal_ip, tenant=parts[4].strip() if len(parts) > 4 else "",
            vcpus=int(parts[5]) if len(parts) > 5 and parts[5].strip().isdigit() else 4,
            memory=int(parts[6]) if len(parts) > 6 and parts[6].strip().isdigit() else 8192,
            disk=parts[7].strip() if len(parts) > 7 else "",
            access_url=parts[8].strip() if len(parts) > 8 else "",
        )
        db.session.add(vm)
        success_count += 1

    db.session.commit()
    return success(data={
        "success_count": success_count,
        "fail_count": fail_count,
        "errors": errors,
    }, msg="导入完成")


def _import_excel(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return error(msg="缺少 openpyxl 依赖，请运行 pip install openpyxl")

    wb = load_workbook(file_obj)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    success_count = 0
    fail_count = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row[0]:
            continue

        name = str(row[0]).strip()
        cluster = str(row[1]).strip() if len(row) > 1 else ""
        external_ip = str(row[2]).strip() if len(row) > 2 else ""
        internal_ip = str(row[3]).strip() if len(row) > 3 else ""

        if CmdbVM.query.filter_by(name=name, deleted=0).first():
            fail_count += 1
            errors.append(f"第{i}行：名称 '{name}' 已存在")
            continue

        vm = CmdbVM(
            name=name, cluster=cluster, external_ip=external_ip,
            internal_ip=internal_ip,
            tenant=str(row[4]).strip() if len(row) > 4 else "",
            vcpus=int(row[5]) if len(row) > 5 and isinstance(row[5], (int, float)) else 4,
            memory=int(row[6]) if len(row) > 6 and isinstance(row[6], (int, float)) else 8192,
            disk=str(row[7]).strip() if len(row) > 7 else "",
            access_url=str(row[8]).strip() if len(row) > 8 else "",
        )
        db.session.add(vm)
        success_count += 1

    db.session.commit()
    return success(data={
        "success_count": success_count,
        "fail_count": fail_count,
        "errors": errors,
    }, msg="导入完成")


@cmdb_bp.route("/vms/export", methods=["GET"])
@jwt_required()
def export_vms():
    keywords = request.args.get("keywords", "")
    cluster = request.args.get("cluster", "")
    status = request.args.get("status")
    tenant = request.args.get("tenant", "")

    q = CmdbVM.query.filter_by(deleted=0)
    if keywords:
        kw = f"%{keywords}%"
        q = q.filter(
            db.or_(
                CmdbVM.name.like(kw), CmdbVM.cluster.like(kw),
                CmdbVM.description.like(kw), CmdbVM.external_ip.like(kw),
                CmdbVM.internal_ip.like(kw),
            )
        )
    if cluster:
        q = q.filter_by(cluster=cluster)
    if status is not None:
        q = q.filter_by(status=int(status))
    if tenant:
        q = q.filter_by(tenant=tenant)

    vms = q.all()

    try:
        from openpyxl import Workbook
    except ImportError:
        return error(msg="缺少 openpyxl 依赖")

    wb = Workbook()
    ws = wb.active
    ws.title = "虚拟机列表"
    headers = ["名称", "集群", "外部IP", "内部IP", "描述", "状态", "租户", "VCPUS", "内存(MB)", "硬盘", "访问URL"]
    ws.append(headers)

    for vm in vms:
        ws.append([
            vm.name, vm.cluster, vm.external_ip, vm.internal_ip,
            vm.description or "", "在线" if vm.status == 1 else "离线",
            vm.tenant, vm.vcpus, vm.memory, vm.disk or "", vm.access_url or "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="cmdb_vms.xlsx",
    )


@cmdb_bp.route("/clusters", methods=["GET"])
@jwt_required()
def list_clusters():
    clusters = (
        db.session.query(CmdbVM.cluster)
        .filter_by(deleted=0)
        .distinct()
        .order_by(CmdbVM.cluster)
        .all()
    )
    data = [{"label": c[0], "value": c[0]} for c in clusters]
    return success(data=data)


@cmdb_bp.route("/tenants", methods=["GET"])
@jwt_required()
def list_tenants():
    tenants = (
        db.session.query(CmdbVM.tenant)
        .filter_by(deleted=0)
        .distinct()
        .order_by(CmdbVM.tenant)
        .all()
    )
    data = [{"label": t[0], "value": t[0]} for t in tenants]
    return success(data=data)
